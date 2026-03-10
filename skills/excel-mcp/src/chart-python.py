#!/usr/bin/env python
"""
Excel 图表生成器 - 使用 matplotlib 生成图表并插入 Excel

依赖：
pip install matplotlib openpyxl pandas

使用示例：
python chart-python.py --input data.xlsx --sheet Sheet1 --range A1:C10 --type column --title "销售趋势" --output result.xlsx --position E1
"""

import argparse
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import tempfile

def create_chart(data, chart_type, title, x_label=None, y_label=None, width=10, height=6):
    """创建图表并返回图片路径"""
    
    # 创建图表
    fig, ax = plt.subplots(figsize=(width, height), dpi=100)
    
    # 中文字体支持
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    
    if chart_type in ['column', 'bar']:
        # 柱状图/条形图
        x = range(len(data['labels']))
        for i, (label, dataset) in enumerate(data['datasets'].items()):
            if chart_type == 'column':
                ax.bar(x, dataset, label=label, alpha=0.7)
            else:
                ax.barh(x, dataset, label=label, alpha=0.7)
        ax.set_xticks(x)
        ax.set_xticklabels(data['labels'])
        
    elif chart_type == 'line':
        # 折线图
        for label, dataset in data['datasets'].items():
            ax.plot(data['labels'], dataset, marker='o', label=label, linewidth=2)
        ax.legend()
        
    elif chart_type == 'pie':
        # 饼图
        labels = list(data['datasets'].keys())
        values = [sum(v) for v in data['datasets'].values()]
        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
        
    elif chart_type == 'scatter':
        # 散点图
        for label, dataset in data['datasets'].items():
            ax.scatter(data['labels'][:len(dataset)], dataset, label=label, alpha=0.6)
            
    elif chart_type == 'area':
        # 面积图
        for label, dataset in data['datasets'].items():
            ax.fill_between(range(len(dataset)), dataset, alpha=0.3, label=label)
        ax.set_xticks(range(len(data['labels'])))
        ax.set_xticklabels(data['labels'])
    
    # 设置标题和标签
    if title:
        ax.set_title(title, fontsize=14, fontweight='bold')
    if x_label:
        ax.set_xlabel(x_label)
    if y_label:
        ax.set_ylabel(y_label)
    
    # 旋转 x 轴标签
    if len(data['labels']) > 5:
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')
    
    # 添加网格
    ax.grid(True, alpha=0.3)
    
    # 添加图例
    if len(data['datasets']) > 1:
        ax.legend()
    
    # 自动调整布局
    plt.tight_layout()
    
    # 保存图片
    temp_path = os.path.join(tempfile.gettempdir(), f'chart_{os.getpid()}.png')
    plt.savefig(temp_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    
    return temp_path

def read_excel_data(file_path, sheet_name, data_range):
    """从 Excel 读取数据"""
    
    # 解析范围（如 'A1:C10'）
    start_cell, end_cell = data_range.split(':')
    
    # 读取数据
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # 转换为图表数据格式
    labels = df.iloc[1:, 0].tolist()  # 第一列作为标签
    datasets = {}
    
    for col in df.columns[1:]:  # 其他列作为数据集
        datasets[str(col)] = df.iloc[1:, df.columns.get_loc(col)].tolist()
    
    return {
        'labels': labels,
        'datasets': datasets
    }

def insert_chart_to_excel(file_path, sheet_name, position, chart_image_path):
    """将图表图片插入 Excel"""
    
    # 加载 Excel 文件
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    # 解析位置（如 'E1'）
    import re
    match = re.match(r'([A-Z]+)(\d+)', position)
    if match:
        col_letter = match.group(1)
        row_number = int(match.group(2))
        
        # Excel 列字母转数字（A=1, B=2, ...）
        col_number = 0
        for char in col_letter:
            col_number = col_number * 26 + (ord(char) - ord('A') + 1)
        
        # 插入图片
        img = Image(chart_image_path)
        
        # 调整图片大小
        img.width = 600
        img.height = 400
        
        # 计算单元格位置
        from openpyxl.utils import get_column_letter
        ws.add_image(img, f'{col_letter}{row_number}')
    
    # 保存文件
    wb.save(file_path)
    print(f"✅ 图表已插入到 {sheet_name}!{position}")

def main():
    parser = argparse.ArgumentParser(description='Excel 图表生成器')
    parser.add_argument('--input', required=True, help='输入 Excel 文件路径')
    parser.add_argument('--sheet', default='Sheet1', help='工作表名称')
    parser.add_argument('--range', required=True, help='数据范围（如 A1:C10）')
    parser.add_argument('--type', required=True, 
                       choices=['column', 'bar', 'line', 'pie', 'scatter', 'area'],
                       help='图表类型')
    parser.add_argument('--title', default='', help='图表标题')
    parser.add_argument('--xlabel', default='', help='X 轴标签')
    parser.add_argument('--ylabel', default='', help='Y 轴标签')
    parser.add_argument('--output', help='输出文件路径（默认覆盖原文件）')
    parser.add_argument('--position', default='E1', help='图表插入位置')
    parser.add_argument('--width', type=int, default=10, help='图表宽度（英寸）')
    parser.add_argument('--height', type=int, default=6, help='图表高度（英寸）')
    
    args = parser.parse_args()
    
    # 输出文件路径
    output_path = args.output or args.input
    
    # 设置 UTF-8 编码输出（Windows 兼容）
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    print(f"[Chart] 生成图表：{args.type}")
    print(f"   数据范围：{args.range}")
    print(f"   标题：{args.title}")
    
    # 1. 读取数据
    data = read_excel_data(args.input, args.sheet, args.range)
    print(f"   标签数：{len(data['labels'])}")
    print(f"   数据集：{len(data['datasets'])}")
    
    # 2. 生成图表
    chart_path = create_chart(
        data,
        args.type,
        args.title,
        args.xlabel,
        args.ylabel,
        args.width,
        args.height
    )
    print(f"[OK] 图表已生成：{chart_path}")
    
    # 3. 插入 Excel
    insert_chart_to_excel(output_path, args.sheet, args.position, chart_path)
    
    # 4. 清理临时文件
    os.remove(chart_path)
    
    print(f"[DONE] 完成！文件已保存：{output_path}")

if __name__ == '__main__':
    main()
